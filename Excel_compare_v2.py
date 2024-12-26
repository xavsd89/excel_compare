import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import streamlit as st
import tempfile

def highlight_col(file_path):
    """
    Highlights differences in an Excel file based on merge status.
    """

    # Load the workbook from the given file path
    wb = load_workbook(file_path)
    # Access the 'Differences' sheet
    ws_diff = wb['Differences']

    # Define fill colors for highlighting cells
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    pink_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Convert ws_diff.columns to a list to safely use len()
    columns = list(ws_diff.columns)
    num_cols = len(columns) - 1  # excluding the last column (merge status)

    # Iterate over rows in the 'Differences' sheet, starting from the second row
    for row in ws_diff.iter_rows(min_row=2, max_col=ws_diff.max_column):
        # Get the merge status from the last column of the row
        merge_output = row[-1].value

        # Highlight the entire row based on the merge status
        for cell in row:
            if merge_output == 'Source Only':
                cell.fill = yellow_fill
            elif merge_output == 'Target Only':
                cell.fill = pink_fill

    # Iterate over rows to compare cell values
    for row in ws_diff.iter_rows(min_row=2, max_col=num_cols):
        # Get the merge status from the last column of the row
        merge_output = row[-1].value
        if merge_output in ['Source Only', 'Target Only']:
            # Compare cells from two halves of the row
            for i in range(num_cols // 2):
                source_cell = row[i]
                target_cell = row[i + num_cols // 2]

                # Highlight cells in red if they are different
                if source_cell.value != target_cell.value:
                    source_cell.fill = red_fill
                    target_cell.fill = red_fill

    # Save the workbook with the applied highlights
    wb.save(file_path)

def main():
    """
    Main function to run the Streamlit application for comparing and merging Excel files.
    """
    
    # Display functional steps at the top for user guidance
    st.title('Excel Compare Tool')
    st.subheader("How it Works:")
    st.write("""
    1. **Upload the Files**: Upload the two Excel files you want to compare.
    2. **Compare the Files**: Click on "Compare Excel Files" to see differences between the two files.
    3. **Download the Results**: Download the resulting Excel file that has generated a 'Differences' tab showing the differences highlighted.
    4. **Color Coding**: 
        - Yellow: Data rows that only exists in the SOURCE file (1st excel).
        - Pink: Data rows that only exists in the TARGET file (2nd excel).
        - Red: Data rows that exists in both files but some Data Attributes has difference between the two files.
    """)

    # Create file uploader widgets for uploading Excel files for comparison
    uploaded_source_file = st.file_uploader("Upload SOURCE Excel File for comparison", type=['xlsx'], key='source')
    uploaded_target_file = st.file_uploader("Upload TARGET Excel File for comparison", type=['xlsx'], key='target')

    # Check if the "Compare Excel Files" button is clicked
    if st.button("Compare Excel Files"):
        if uploaded_source_file and uploaded_target_file:
            try:
                # Read the uploaded Excel files into DataFrames
                source_file = pd.read_excel(uploaded_source_file)
                target_file = pd.read_excel(uploaded_target_file)

                # Strip whitespace from string values in both DataFrames
                source_file = source_file.applymap(lambda x: str(x).strip() if isinstance(x, str) else x)
                target_file = target_file.applymap(lambda x: str(x).strip() if isinstance(x, str) else x)

                # Perform an outer merge to find differences between source and target files
                diff = source_file.merge(target_file, indicator=True, how='outer')
                # Replace merge indicator values with custom labels
                diff['_merge'].replace({'left_only': 'Source Only', 'right_only': 'Target Only'}, inplace=True)
                # Rename the merge indicator column to 'Merge_Output'
                diff.rename(columns={'_merge': 'Merge_Output'}, inplace=True)
                # Filter out rows where both source and target have the same data
                diff = diff[diff['Merge_Output'] != 'both']

                # Define the path for the differences file
                diff_file_path = 'diff_output.xlsx'

                if not diff.empty:
                    # Write source, target, and differences to the Excel file
                    with pd.ExcelWriter(diff_file_path) as writer:
                        source_file.to_excel(writer, sheet_name='Source', index=False)
                        target_file.to_excel(writer, sheet_name='Target', index=False)
                        diff.to_excel(writer, sheet_name='Differences', index=False)
                    
                    # Apply highlighting to the differences sheet
                    highlight_col(diff_file_path)

                    # Notify the user of found differences and provide download link
                    st.write(f"Found {len(diff)} differences.")
                    with open(diff_file_path, "rb") as f:
                        st.download_button("Download Differences File", f, file_name="diff_output.xlsx")
                else:
                    # Create an Excel file with no differences
                    with pd.ExcelWriter(diff_file_path) as writer:
                        source_file.to_excel(writer, sheet_name='Source', index=False)
                        target_file.to_excel(writer, sheet_name='Target', index=False)
                        pd.DataFrame(columns=source_file.columns).to_excel(writer, sheet_name='Differences', index=False)
                        writer.sheets['Differences'].write(0, 0, "No difference found between source and target files.")
                    
                    # Notify the user that no differences were found and provide download link
                    st.write("No differences found.")
                    with open(diff_file_path, "rb") as f:
                        st.download_button("Download Differences File", f, file_name="diff_output.xlsx")

            except Exception as e:
                # Display error message if an exception occurs
                st.error(f"An error occurred: {e}")
        else:
            # Display error message if both files are not uploaded
            st.error("Please upload both Excel files.")

if __name__ == "__main__":
    # Run the main function to start the Streamlit application
    main()
